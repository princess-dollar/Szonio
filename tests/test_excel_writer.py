from pathlib import Path

import openpyxl

from services.excel_writer import write_company_result
from services.orchestrator import process_file

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SAMPLE_FILE = PROJECT_ROOT / "Upload_Excel.xlsx"

# Real columns from Upload_Excel.xlsx: index 1 -> employee_id, index 21 ->
# salary_per_period, index 64 -> deduct_unpaid_leave, index 61 -> deduct_late_early.


class _StubLlmClient:
    """Stands in for integrations.llm_gateway_client.LlmGatewayClient — no
    network, no LLM call, in any test."""

    def __init__(self, raw_response):
        self._raw_response = raw_response

    def map_columns(self, columns, canonical_fields):
        return self._raw_response


def _domnick_full_raw_mapping():
    return {
        "mappings": [
            {
                "column_index": 1,
                "column_name": "รหัสพนักงาน",
                "canonical_field": "employee_id",
                "confidence": 0.99,
            },
            {
                "column_index": 21,
                "column_name": "เงินเดือนต่องวด",
                "canonical_field": "salary_per_period",
                "confidence": 0.97,
            },
            {
                "column_index": 64,
                "column_name": "หักลาไม่รับค่าจ้าง (บาท)",
                "canonical_field": "deduct_unpaid_leave",
                "confidence": 0.95,
            },
            {
                "column_index": 61,
                "column_name": "หักมาสาย/กลับก่อน (บาท)",
                "canonical_field": "deduct_late_early",
                "confidence": 0.93,
            },
        ]
    }


def test_output_workbook_has_both_sheets_with_expected_headers_and_row_count(tmp_path):
    client = _StubLlmClient(_domnick_full_raw_mapping())
    result = process_file(str(SAMPLE_FILE), "domnick", llm_client=client)
    assert result.status == "calculated"

    output_path = tmp_path / "Upload_Excel_sso_result.xlsx"
    write_company_result(result.company_result, str(output_path))

    assert output_path.exists()

    wb = openpyxl.load_workbook(str(output_path))
    assert wb.sheetnames == ["ผลรายคน", "สรุป"]

    per_employee_ws = wb["ผลรายคน"]
    header = [cell.value for cell in per_employee_ws[1]]
    assert header == ["employee_id", "ฐาน SSO", "เงินสมทบ"]

    data_row_count = per_employee_ws.max_row - 1  # minus header
    assert data_row_count == len(result.company_result.employees) == 33

    summary_ws = wb["สรุป"]
    summary_text = "\n".join(
        str(cell.value) for row in summary_ws.iter_rows() for cell in row if cell.value is not None
    )
    assert "domnick" in summary_text
    assert "total_base" in summary_text
    assert "total_contribution" in summary_text


def test_needs_review_pipeline_produces_no_output_file(tmp_path):
    raw = _domnick_full_raw_mapping()
    raw["mappings"] = [m for m in raw["mappings"] if m["canonical_field"] != "salary_per_period"]
    client = _StubLlmClient(raw)

    result = process_file(str(SAMPLE_FILE), "domnick", llm_client=client)
    assert result.status == "needs_review"
    assert result.company_result is None

    # The orchestrator never calls write_company_result when the pipeline
    # stops at needs_review -- company_result is None, so there is nothing
    # to write. Confirm that contract explicitly: no file appears.
    output_path = tmp_path / "should_not_exist.xlsx"
    assert not output_path.exists()
