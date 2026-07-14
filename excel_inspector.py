"""Phase 1: inspect a payroll Excel export's real structure.

Never hardcode a header row or column layout here — both are detected from
the file itself (see .claude/skills/sso-calc/SKILL.md, rule 5).
"""

import re
from collections import Counter
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from models import ColumnInfo, WorkbookMetadata

PII_COLUMN_NAMES = {"ชื่อ-นามสกุล", "เลขบัญชี", "บัญชีธนาคาร"}
GROUND_TRUTH_NAMES = {"BASE SSO", "CAL SSO", "CHECK", "BASE TAX", "TAX"}
TOTAL_ROW_SENTINELS = {"total", "รวม", "รวมทั้งหมด", "grand total"}

_DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
_HEADER_TEXT_RATIO = 0.5
_MAX_HEADER_SCAN_ROWS = 10


def _is_probable_header_row(values: tuple[Any, ...]) -> bool:
    """The index row is ints, the group row is mostly empty — only the real
    header row is mostly non-empty text, so this alone separates all three."""
    if not values:
        return False
    text_cells = sum(1 for v in values if isinstance(v, str) and v.strip() != "")
    return (text_cells / len(values)) > _HEADER_TEXT_RATIO


def _detect_header_row(ws: Worksheet) -> int:
    for row_idx in range(1, _MAX_HEADER_SCAN_ROWS + 1):
        try:
            values = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        except StopIteration:
            break
        if _is_probable_header_row(values):
            return row_idx
    raise ValueError(
        f"Could not auto-detect a header row in the first {_MAX_HEADER_SCAN_ROWS} rows"
    )


def _forward_fill_groups(values: tuple[Any, ...]) -> list[Optional[str]]:
    groups: list[Optional[str]] = []
    current: Optional[str] = None
    for v in values:
        if v is not None and str(v).strip() != "":
            current = str(v).strip()
        groups.append(current)
    return groups


def _classify_value(value: Any) -> str:
    if isinstance(value, bool):
        return "text"
    if isinstance(value, (int, float)):
        return "numeric"
    if isinstance(value, str) and _DATE_RE.match(value.strip()):
        return "date"
    return "text"


def _infer_column_type(values: list[Any]) -> str:
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "text"
    counts = Counter(_classify_value(v) for v in non_empty)
    return counts.most_common(1)[0][0]


def _is_total_row(first_cell: Any) -> bool:
    if first_cell is None:
        return False
    return str(first_cell).strip().casefold() in TOTAL_ROW_SENTINELS


def _mask_pii(name: str, value: Any) -> Any:
    if name in PII_COLUMN_NAMES and value is not None:
        return "<masked>"
    return value


def inspect_workbook(path: str, sample_size: int = 5) -> WorkbookMetadata:
    """Read a payroll Excel export and report its real structure.

    Trailing summary/total rows (first cell matching a known sentinel like
    "Total"/"รวม") are excluded from row_count and sample_rows — they are not
    employee data.
    """
    if not (3 <= sample_size <= 10):
        raise ValueError("sample_size must be between 3 and 10")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        header_row_idx = _detect_header_row(ws)
        header_values = next(
            ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
        )

        if header_row_idx > 1:
            group_values = next(
                ws.iter_rows(
                    min_row=header_row_idx - 1, max_row=header_row_idx - 1, values_only=True
                )
            )
            groups = _forward_fill_groups(group_values)
        else:
            groups = [None] * len(header_values)

        total_columns = len(header_values)

        data_rows = [
            row
            for row in ws.iter_rows(
                min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True
            )
            if not _is_total_row(row[0] if row else None)
        ]

        columns: list[ColumnInfo] = []
        ground_truth: list[ColumnInfo] = []
        for col_pos in range(total_columns):
            idx = col_pos + 1
            raw_name = header_values[col_pos]
            name = str(raw_name).strip() if raw_name is not None else ""
            column_values = [row[col_pos] if col_pos < len(row) else None for row in data_rows]
            info = ColumnInfo(
                index=idx,
                group=groups[col_pos],
                name=name,
                inferred_type=_infer_column_type(column_values),
            )
            if name in GROUND_TRUTH_NAMES:
                ground_truth.append(info)
            else:
                columns.append(info)

        normal_indexes = {c.index for c in columns}
        sample_rows: list[dict[int, Any]] = []
        for row in data_rows[:sample_size]:
            record: dict[int, Any] = {}
            for col_pos, value in enumerate(row):
                idx = col_pos + 1
                if idx not in normal_indexes:
                    continue
                raw_name = header_values[col_pos]
                name = str(raw_name).strip() if raw_name is not None else ""
                record[idx] = _mask_pii(name, value)
            sample_rows.append(record)

        return WorkbookMetadata(
            sheet_name=ws.title,
            header_row_index=header_row_idx,
            columns=columns,
            sample_rows=sample_rows,
            row_count=len(data_rows),
            ground_truth=ground_truth,
        )
    finally:
        wb.close()
