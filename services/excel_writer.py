"""Phase 5's output writer: a brand-new .xlsx with two sheets summarizing
one company's calculated SSO results.

Deliberately no Excel formulas here, unlike a typical financial-model
workbook: every number is a value Python already computed with Decimal
precision in Phase 4 (SKILL.md rule 4 — Python owns all business logic and
number crunching). Recomputing them via spreadsheet formulas would hand
that ownership back to Excel, which this project's architecture forbids.

Never touches the source workbook — its 3-row header is fragile; a fresh
file is safer (openpyxl always creates a new Workbook here).
"""

from typing import Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from services.result_builder import CompanyResult

_FONT_NAME = "Arial"
_MONEY_FORMAT = "#,##0.00"
_BOLD = Font(name=_FONT_NAME, bold=True)
_REGULAR = Font(name=_FONT_NAME)


def write_company_result(company_result: CompanyResult, output_path: str) -> None:
    wb = openpyxl.Workbook()

    per_employee_ws = wb.active
    _write_per_employee_sheet(per_employee_ws, company_result)
    _write_summary_sheet(wb.create_sheet("สรุป"), company_result)

    wb.save(output_path)


def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _BOLD


def _write_per_employee_sheet(ws: Worksheet, company_result: CompanyResult) -> None:
    ws.title = "ผลรายคน"
    _write_header_row(ws, ["employee_id", "ฐาน SSO", "เงินสมทบ"])

    for employee in company_result.employees:
        ws.append([employee.employee_id, employee.base, employee.contribution])

    for row in ws.iter_rows(min_row=2):
        row[0].font = _REGULAR
        for cell in row[1:]:
            cell.font = _REGULAR
            cell.number_format = _MONEY_FORMAT

    for column_letter, width in (("A", 16), ("B", 14), ("C", 14)):
        ws.column_dimensions[column_letter].width = width


def _write_kv_row(ws: Worksheet, label: str, value) -> None:
    ws.append([label, value])
    ws[f"A{ws.max_row}"].font = _REGULAR


def _write_section_title(ws: Worksheet, title: str) -> None:
    ws.append([title])
    ws[f"A{ws.max_row}"].font = _BOLD


def _write_summary_sheet(ws: Worksheet, company_result: CompanyResult) -> None:
    ws.title = "สรุป"
    audit = company_result.audit

    _write_section_title(ws, "ข้อมูลบริษัท")
    _write_kv_row(ws, "บริษัท (company_id)", audit.company_id)
    _write_kv_row(ws, "ชื่อบริษัท (display_name)", audit.display_name)
    _write_kv_row(ws, "เวอร์ชัน config (config_version)", audit.config_version)
    ws.append([])

    _write_section_title(ws, "กฎเงินสมทบประกันสังคม (sso_rule)")
    _write_kv_row(ws, "อัตราเงินสมทบ (rate)", audit.sso_rate)
    ws[f"B{ws.max_row}"].number_format = "0.00%"
    _write_kv_row(ws, "เพดานฐานเงินเดือน (ceiling)", audit.sso_ceiling)
    ws[f"B{ws.max_row}"].number_format = _MONEY_FORMAT
    ws.append([])

    _write_section_title(ws, "สรุปยอดรวม (totals)")
    _write_kv_row(ws, "จำนวนพนักงานที่คำนวณสำเร็จ", len(company_result.employees))
    _write_kv_row(ws, "รวมฐาน SSO ทั้งหมด (total_base)", company_result.total_base)
    ws[f"B{ws.max_row}"].number_format = _MONEY_FORMAT
    _write_kv_row(ws, "รวมเงินสมทบทั้งหมด (total_contribution)", company_result.total_contribution)
    ws[f"B{ws.max_row}"].number_format = _MONEY_FORMAT
    _write_kv_row(ws, "จำนวนแถวที่มีปัญหา (error_rows)", len(company_result.error_rows))
    ws.append([])

    _write_section_title(ws, "สูตรคำนวณและการแมปคอลัมน์ที่ใช้ (formula components + column mapping)")
    _write_header_row(ws, ["canonical_field", "sign", "column_index", "column_name"])
    for item in audit.column_mapping:
        ws.append(
            [
                item.canonical_field,
                item.sign if item.sign is not None else "",
                item.column_index if item.column_index is not None else "(ไม่ได้แมป)",
                item.column_name if item.column_name is not None else "(ไม่ได้แมป)",
            ]
        )
    ws.append([])

    _write_section_title(ws, "แถวที่มีปัญหา (error rows)")
    if company_result.error_rows:
        _write_header_row(ws, ["row_number", "employee_id", "reason"])
        for error in company_result.error_rows:
            ws.append([error.row_number, error.employee_id or "", error.reason])
    else:
        ws.append(["(ไม่มี)"])
        ws[f"A{ws.max_row}"].font = _REGULAR

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
